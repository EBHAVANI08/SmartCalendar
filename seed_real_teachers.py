import openpyxl
import sqlite3
import json
import random
import re

excel_path = r"C:\Users\Bhavani\Downloads\Teacher_Allocation_Typed.xlsx"
db_path = r"C:\Users\Bhavani\OneDrive - KAM Global for Digital & AI Media Solutions Private Limited\kiran\Smart_Calendar\prisma\db\custom.db"

wb = openpyxl.load_workbook(excel_path, data_only=True)

# 1. Parse Class Teachers sheet
class_teachers = {}
ws_ct = wb['Class Teachers']
for r in range(1, ws_ct.max_row + 1):
    c1 = ws_ct.cell(r, 1).value
    c2 = ws_ct.cell(r, 2).value
    if c1 and c2:
        class_key = str(c1).strip()
        name = str(c2).strip()
        class_teachers[class_key] = name

# 2. Parse subject allotment sheet
ws_sa = wb['subject allotment']
headers = [str(ws_sa.cell(1, c).value or '').strip() for c in range(1, ws_sa.max_column + 1)]

teachers_map = {} # teacher_name -> {subjects: set(), grades: set(), is_ct_for: set()}
grade_subject_teacher = {} # (grade_str, subject) -> teacher_name

def add_teacher_info(teacher_name, subject, grade_str, is_class_teacher=False):
    if not teacher_name or str(teacher_name).upper() in ['NO', 'NONE', '']:
        return
    tname = str(teacher_name).strip()
    if tname not in teachers_map:
        teachers_map[tname] = {'subjects': set(), 'grades': set(), 'is_ct_for': set()}
    if subject and subject != 'Class Teacher':
        teachers_map[tname]['subjects'].add(subject)
    if grade_str:
        teachers_map[tname]['grades'].add(grade_str)
    if is_class_teacher:
        teachers_map[tname]['is_ct_for'].add(grade_str)

# Record Class Teachers
for grade_str, teacher_name in class_teachers.items():
    add_teacher_info(teacher_name, 'Class Teacher', grade_str, is_class_teacher=True)

# Record Subject Teachers
for r in range(2, ws_sa.max_row + 1):
    grade_val = str(ws_sa.cell(r, 1).value or '').strip()
    if not grade_val:
        continue
    for c in range(2, len(headers)):
        subj = headers[c]
        tval = ws_sa.cell(r, c).value
        if tval and str(tval).strip().upper() not in ['NO', 'NONE', '']:
            tname = str(tval).strip()
            add_teacher_info(tname, subj, grade_val)
            grade_subject_teacher[(grade_val, subj)] = tname

print(f"Total Unique Real Teachers parsed: {len(teachers_map)}")

# Connect to database
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Clean old records for Sunrise Public School
school_id = "sch_sunrise_001"
cursor.execute("DELETE FROM Schedule WHERE schoolId = ?", (school_id,))
cursor.execute("DELETE FROM Teacher WHERE schoolId = ?", (school_id,))

# Insert real teachers into database
teacher_db_ids = {} # teacher_name -> db_id
days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

created_count = 0
for idx, (tname, info) in enumerate(teachers_map.items(), start=1):
    db_id = f"tch_real_{idx:03d}"
    teacher_db_ids[tname] = db_id
    
    # Generate clean email
    clean_name = re.sub(r'[^a-zA-Z]', '', tname.lower())
    email = f"{clean_name}@sunrisepublic.edu"
    
    subjects_list = list(info['subjects'])
    primary_subject = subjects_list[0] if subjects_list else "General"
    grades_list = list(info['grades'])
    
    # Append Class Teacher tag to name if applicable
    ct_info = f" (Class Teacher: {', '.join(info['is_ct_for'])})" if info['is_ct_for'] else ""
    display_name = f"{tname}{ct_info}"
    
    cursor.execute("""
        INSERT INTO Teacher (id, name, email, subject, grades, availability, schoolId, createdAt, updatedAt)
        VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
    """, (db_id, display_name, email, primary_subject, json.dumps(grades_list), 'available', school_id))
    created_count += 1

print(f"Inserted {created_count} real teachers into Sunrise Public School.")

# Parse grades and sections cleanly
def parse_grade_section(gstr):
    gstr = gstr.strip()
    match1 = re.match(r'^(\d+)-?\s*(.+)$', gstr)
    if match1:
        return f"Grade {match1.group(1)}", match1.group(2).capitalize()
    
    roman_map = {'I': '1', 'II': '2', 'III': '3', 'IV': '4', 'V': '5', 'VI': '6', 'VII': '7', 'VIII': '8', 'IX': '9', 'X': '10'}
    match2 = re.match(r'^(I|II|III|IV|V|VI|VII|VIII|IX|X)\s*(.+)$', gstr, re.IGNORECASE)
    if match2:
        num = roman_map.get(match2.group(1).upper(), match2.group(1))
        return f"Grade {num}", match2.group(2).capitalize()
    
    return "Grade 1", gstr

# Build schedules for all grades based on Excel assignments
schedules_created = 0
unique_classes = set([g for (g, s) in grade_subject_teacher.keys()] + list(class_teachers.keys()))

for gstr in unique_classes:
    grade_name, section_name = parse_grade_section(gstr)
    
    # Find all subject-teacher mappings for this class
    class_subjects = [(s, t) for (g, s), t in grade_subject_teacher.items() if g == gstr]
    if not class_subjects:
        continue
        
    for day in days:
        for period in range(1, 9): # 8 periods per day
            subj, tname = class_subjects[(period - 1) % len(class_subjects)]
            tid = teacher_db_ids.get(tname)
            
            sch_id = f"sch_real_{schedules_created+1:05d}"
            expected_topic = f"{subj} Unit {period} - Chapter Exercises"
            
            start_hour = 8 + (period - 1)
            start_time = f"{start_hour:02d}:00"
            end_time = f"{start_hour:02d}:45"
            
            cursor.execute("""
                INSERT INTO Schedule (id, grade, section, day, period, subject, topic, startTime, endTime, teacherId, schoolId, createdAt, updatedAt)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
            """, (sch_id, grade_name, section_name, day, period, subj, expected_topic, start_time, end_time, tid, school_id))
            schedules_created += 1

conn.commit()
conn.close()

print(f"Successfully seeded {schedules_created} schedule slots for Sunrise Public School with real teacher allocations!")
