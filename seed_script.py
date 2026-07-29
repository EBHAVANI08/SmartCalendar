import sqlite3
import openpyxl
import os
import uuid
from datetime import datetime

db_path = r'C:\Users\Bhavani\OneDrive - KAM Global for Digital & AI Media Solutions Private Limited\kiran\Smart_Calendar\prisma\db\custom.db'
excel_path = r'C:\Users\Bhavani\Downloads\Teacher_Allocation_Dummy_Data.xlsx'

conn = sqlite3.connect(db_path)
cursor = conn.cursor()

now_str = datetime.utcnow().isoformat() + 'Z'

# 1. Create Schools
sunrise_school_id = 'sch_sunrise_001'
greenwood_school_id = 'sch_greenwood_002'

schools = [
    (sunrise_school_id, 'Sunrise Public School', 'SUNRISE', 'admin@sunrisepublic.edu', 'school123', now_str, now_str),
    (greenwood_school_id, 'Greenwood High School', 'GREENWOOD', 'admin@greenwoodhigh.edu', 'school123', now_str, now_str)
]

for sch in schools:
    cursor.execute('''
        INSERT OR REPLACE INTO School (id, name, code, email, password, createdAt, updatedAt)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', sch)

print("Inserted/updated 2 demo schools: Sunrise Public School & Greenwood High School")

# 2. Parse Excel
wb = openpyxl.load_workbook(excel_path)
ws_allotment = wb['subject allotment']
ws_class_teachers = wb['Class Teachers']

headers = [cell for cell in next(ws_allotment.iter_rows(values_only=True))]
subjects_list = headers[1:-1] # Skip Grade and School Name

teacher_name_to_id = {}
teacher_subject_map = {}

# Read rows
row_idx = 1
for row in ws_allotment.iter_rows(min_row=2, values_only=True):
    grade_sec = str(row[0]) if row[0] else ''
    if not grade_sec:
        continue
    for col_i, sub in enumerate(subjects_list, start=1):
        t_name = str(row[col_i]).strip() if row[col_i] else ''
        if t_name and t_name.upper() not in ['NO', 'NONE', '']:
            if t_name not in teacher_name_to_id:
                t_id = f"tch_sun_{len(teacher_name_to_id) + 1:03d}"
                teacher_name_to_id[t_name] = t_id
                teacher_subject_map[t_name] = sub

print(f"Parsed {len(teacher_name_to_id)} unique teachers for Sunrise Public School.")

# Insert Teachers for Sunrise Public School
for t_name, t_id in teacher_name_to_id.items():
    clean_num = ''.join(filter(str.isdigit, t_name)) or '01'
    email = f"teacher{clean_num}.sunrise@smartcal.com"
    subject = teacher_subject_map.get(t_name, 'General')
    
    cursor.execute('''
        INSERT OR REPLACE INTO Teacher (id, name, email, password, phone, subject, grades, availability, role, schoolId, createdAt, updatedAt)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        t_id,
        t_name,
        email,
        'teacher123',
        '+91 98765 43210',
        subject,
        '["Grade 1","Grade 2","Grade 3"]',
        '[]',
        'teacher',
        sunrise_school_id,
        now_str,
        now_str
    ))

# Also add 5 sample dummy teachers for Greenwood High School
greenwood_teachers = [
    ('tch_gw_001', 'Dr. Alan Grant (Greenwood)', 'grant@greenwoodhigh.edu', 'Science', greenwood_school_id),
    ('tch_gw_002', 'Prof. Sarah Connor (Greenwood)', 'sarah@greenwoodhigh.edu', 'Mathematics', greenwood_school_id),
    ('tch_gw_003', 'Mr. Bruce Wayne (Greenwood)', 'bruce@greenwoodhigh.edu', 'Physical Education', greenwood_school_id),
    ('tch_gw_004', 'Ms. Diana Prince (Greenwood)', 'diana@greenwoodhigh.edu', 'Social Studies', greenwood_school_id),
    ('tch_gw_005', 'Mr. Clark Kent (Greenwood)', 'clark@greenwoodhigh.edu', 'English', greenwood_school_id),
]

for t in greenwood_teachers:
    cursor.execute('''
        INSERT OR REPLACE INTO Teacher (id, name, email, password, phone, subject, grades, availability, role, schoolId, createdAt, updatedAt)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        t[0], t[1], t[2], 'teacher123', '+91 99999 88888', t[3], '["Grade 9","Grade 10"]', '[]', 'teacher', t[4], now_str, now_str
    ))

# 3. Create Schedules for Sunrise Public School
days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
time_slots = [
    ("08:00", "08:45"),
    ("08:45", "09:30"),
    ("09:30", "10:15"),
    ("10:30", "11:15"),
    ("11:15", "12:00"),
    ("12:45", "13:30"),
    ("13:30", "14:15"),
    ("14:15", "15:00")
]

schedule_count = 0
for row in ws_allotment.iter_rows(min_row=2, values_only=True):
    grade_sec = str(row[0]).strip() if row[0] else ''
    if not grade_sec:
        continue
    
    parts = grade_sec.split(' ', 1)
    grade = parts[0]
    section = parts[1] if len(parts) > 1 else 'A'

    # Get valid subjects & teachers for this grade/sec
    assigned_pairs = []
    for col_i, sub in enumerate(subjects_list, start=1):
        t_name = str(row[col_i]).strip() if row[col_i] else ''
        if t_name and t_name.upper() not in ['NO', 'NONE', '']:
            assigned_pairs.append((sub, teacher_name_to_id[t_name]))

    if not assigned_pairs:
        continue

    # Distribute subjects across 5 days x 8 periods
    for day in days:
        for period_idx in range(1, 9):
            pair = assigned_pairs[(period_idx - 1) % len(assigned_pairs)]
            sub_name, t_id = pair
            start_t, end_t = time_slots[period_idx - 1]
            sch_id = f"sch_{grade}_{section}_{day}_{period_idx}".replace(' ', '_')
            
            cursor.execute('''
                INSERT OR REPLACE INTO Schedule (id, grade, section, day, period, subject, teacherId, schoolId, topic, roomId, startTime, endTime, createdAt, updatedAt)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                sch_id, grade, section, day, period_idx, sub_name, t_id, sunrise_school_id, f"Topic {period_idx}: {sub_name}", f"Room {grade}", start_t, end_t, now_str, now_str
            ))
            schedule_count += 1

# Also create dummy schedules for Greenwood High
for day in days:
    for p in range(1, 6):
        sch_id = f"sch_gw_9A_{day}_{p}"
        cursor.execute('''
            INSERT OR REPLACE INTO Schedule (id, grade, section, day, period, subject, teacherId, schoolId, topic, roomId, startTime, endTime, createdAt, updatedAt)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            sch_id, '9', 'A', day, p, 'Science', 'tch_gw_001', greenwood_school_id, 'Physics Fundamentals', 'Room 101', time_slots[p-1][0], time_slots[p-1][1], now_str, now_str
        ))

conn.commit()
conn.close()

print(f"Successfully seeded database with {len(teacher_name_to_id)} teachers and {schedule_count} schedule entries for Sunrise Public School!")
